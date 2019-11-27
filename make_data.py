from openpyxl import Workbook
import random

# Excelデータを作成する
wb = Workbook()
ws = wb.active
ws.title = 'test_data'

# データの内容を記述
ws['A1'].value = '氏名'
ws['B1'].value = 'HTML/CSS'
ws['C1'].value = 'Rails'
ws['D1'].value = 'JavaScript/jquery'
data_name = ['山下智也','至田昌史','小嶋孝弘','有川恒紀','假屋祥平','吉川翔馬','小林瑞生','石川和良','豊田達哉','矢野翔太','桐谷惇礼','向井健悟','草野真太郎','山本雅夫','福田雄大','谷脇大斗','新井直道','延生光','百済大晃','冨井隆史','金山英祐','原田祐輔','奈須拓実','猪股輝哉','堀優樹','上野和輝','武原徳孝','行広仁','横井祥','畑直樹','西口裕貴','阿部善隆','山本隆之介','柿澤一樹','真鍋康太','門谷晟那','諏訪順一','藤原拓','宮崎文','沢田道宏','末本希望','小林亮介','杉森笑子','末廣拓海','大宮実','阪口玄','植田晃平','神農泰植','北薗忠祥','宮崎祐馬','松川佳世','山角美希','藤田和典','西林正樹']

# 作成したテーブルにランダムなデータを入れていく
for i in range(2, 56):
  ws.cell(row=i, column=1).value = data_name[i - 2]
  _cell1 = ws.cell(row=i, column=2)
  _cell1.value = random.randint(1,100)
  _cell2 = ws.cell(row=i, column=3)
  _cell2.value = random.randint(1,100)
  _cell3 = ws.cell(row=i, column=4)
  _cell3.value = random.randint(1,100)

  
# データを保存する
wb.save('test_data.xlsx')
print('saved.')